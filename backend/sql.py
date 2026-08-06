from constant import db_path, FILES_DIR
import os
import sqlite3, csv
from flask import Flask, render_template, request, Blueprint
from openpyxl import load_workbook, Workbook

database_module = Blueprint("database", __name__, static_folder='./static')

@database_module.route('/database', methods=['POST'])
def database():
    user_id = int(request.form.get('user_id'))
    command = request.form.get('command')
    if (command == 'download'):
        title = '演習問題のダウンロード'
        return render_template('download.html',
                               user_id=user_id,
                               )
    elif (command == 'questions'):
        title = '演習問題の更新'
    elif (command == 'comments'):
        title = 'コメントデータの更新'
    elif (command == 'comment'):
        title = 'コメントデータの更新'
        return render_template('maintenance2.html',
                               user_id=user_id,
                               )
    else:
        return render_template('maintenance.html',
                               user_id=user_id,
                               )
    return render_template('getfile.html',
                           user_id=user_id,
                           title=title,
                           )


def _remove_upload_file(path: str) -> None:
    """Delete temporary upload under static/; ignore if already gone."""
    try:
        if path and os.path.isfile(path):
            os.remove(path)
    except OSError as exc:
        print(f"failed to remove upload file {path}: {exc}")


@database_module.route('/upload', methods=['POST'])
def upload():
    user_id = int(request.form.get('user_id'))
    title = request.form.get('title')
    if title == '演習問題の更新':
        filename = 'QUESTIONSX.XLSX'
    else:
        filename = 'COMMENTSX.XLSX'

    # アップロードしたファイルのオブジェクト
    upfile = request.files.get('upfile', None)
    if upfile is None:
        return render_template('error2.html',
                           user_id=user_id,
                           message='ファイル名が入力されていません。アップロードが失敗しました。'
                           )
    if upfile.filename == '':
        return render_template('error2.html',
                               user_id=user_id,
                               message='ファイル名が入力されていません。アップロードが失敗しました。'
                               )
    saved_path = os.path.join(FILES_DIR, filename)
    upfile.save(saved_path)

    try:
        if title == '演習問題の更新':
            result, value = convertQuestions()
            if result == False:
                if value != -1:
                    return render_template('error3.html',
                                       error_message='識別IDが重なっています。ID=' + str(value)
                                            + '<br>更新前のデータは既に削除されています。'
                                            + '<br>データを修正して再度試みるか、保管している事前のデータをリストアしてください。'
                                           )
                else:
                    return render_template('error3.html',
                                               error_message='EXCELデータの内容に問題があります。'
                                            + '<br>更新前のデータは既に削除されています。'
                                            + '<br>データを修正して再度試みるか、保管している事前のデータをリストアしてください。'
                                               )
        else:
            result, value = convertComments()
            if result == False:
                return render_template('error3.html',
                                       error_message='EXCELデータの内容に問題があります。'
                                            + '<br>更新前のデータは既に削除されています。'
                                            + '<br>データを修正して再度試みるか、保管している事前のデータをリストアしてください。'
                                       )
        return render_template('success.html',
                               user_id=user_id,
                               message='成功しました。'
                               )
    finally:
        # 成功・失敗どちらでも一時 Excel を残さない
        _remove_upload_file(saved_path)

@database_module.route('/download', methods=['POST'])
def download():
    user_id = int(request.form.get('user_id'))

    result = retrieveData()
    if (result == 1):
        return render_template('download2.html',
                           user_id=user_id,
                           message='成功しました。'
                           )
    else:
        return render_template('error2.html',
                           user_id=user_id,
                           message='失敗しました。'
                           )

def convertQuestions():

    # DBに接続してテーブルを作成
    conn = sqlite3.connect(db_path)
    conn.execute("DROP TABLE IF EXISTS knowledge_base")
    conn.execute('''
      CREATE TABLE IF NOT EXISTS knowledge_base (
        number INTEGER PRIMARY KEY,
        category INTEGER,
        level INTEGER,
        q TEXT,
        a1 TEXT,
        a2 TEXT,
        a3 TEXT,
        a4 TEXT,
        cid1 INTEGER,
        cid2 INTEGER,
        cid3 INTEGER,
        cid4 INTEGER,
        flag INTEGER
      )
    ''')

    result, value = read_excel(conn, FILES_DIR + '/QUESTIONSX.XLSX')
    return result, value

def convertComments():
    # DBに接続してテーブルを作成
    conn = sqlite3.connect(db_path)
    conn.execute("DROP TABLE IF EXISTS comments_table")
    conn.execute('''
      CREATE TABLE IF NOT EXISTS comments_table (
        comment_id INTEGER PRIMARY KEY,
        comment TEXT
      )
    ''')

    result, value = read_excel2(conn, FILES_DIR + '/COMMENTSX.XLSX')
    return result, value

def retrieveData():
    # DBに接続してテーブルを作成
    qfile = FILES_DIR + '/questions.xlsx'
    cfile = FILES_DIR + '/comments.xlsx'

    conn = sqlite3.connect(db_path)

    try:
        c = conn.cursor()

        c.execute("SELECT number,level,category,q,a1,a2,a3,a4,cid1,cid2,cid3,cid4,flag FROM knowledge_base")
        rows = c.fetchall()
        print("knowledge_base rows:", len(rows))

        wb1 = Workbook()
        ws1 = wb1.active

        for row in rows:
            ws1.append(row)

        wb1.save(qfile)


        c.execute("SELECT * FROM comments_table")
        rows = c.fetchall()

        wb2 = Workbook()
        ws2 = wb2.active

        for row in rows:
            ws2.append(row)

        wb2.save(cfile)

    except:
        conn.close()
        return -1
    else:
        conn.close()
        return 1

def _is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _as_int(value):
    """Excel cell → int or None (empty allowed). Accepts 101 / 101.0 / '101'."""
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if text == "":
        return None
    return int(float(text))


def _as_text(value):
    if _is_blank(value):
        return None
    if isinstance(value, str):
        return value
    return str(value)


def parse_question_row(row):
    """Parse one Excel row into DB values.

    Accepts 13 columns (A–M) or 14+ columns (N列以降はメモとして読み捨て).
    Returns None for a completely blank row.
    """
    cells = list(row) if row is not None else []
    # Trailing Nones from openpyxl wide rows
    while cells and cells[-1] is None:
        cells.pop()
    if not cells or all(_is_blank(c) for c in cells):
        return None

    if len(cells) < 13:
        cells = cells + [None] * (13 - len(cells))
    # 14列目以降（メモ）は破棄
    number, category, level, q, a1, a2, a3, a4, cid1, cid2, cid3, cid4, flag = cells[:13]

    number_i = _as_int(number)
    if number_i is None:
        raise ValueError("number is required")

    return (
        number_i,
        _as_int(category),
        _as_int(level),
        _as_text(q),
        _as_text(a1),
        _as_text(a2),
        _as_text(a3),
        _as_text(a4),
        _as_int(cid1),
        _as_int(cid2),
        _as_int(cid3),
        _as_int(cid4),
        _as_int(flag),
    )


# EXCELを読んでDBに入れる関数
def read_excel(conn, fname):

    c = conn.cursor()

    wb = load_workbook(fname, data_only=True)
    ws = wb.active  # 1枚目のシート

    number = -1
    try:
        for row in ws.iter_rows(min_row=1, values_only=True):
            parsed = parse_question_row(row)
            if parsed is None:
                continue
            number = parsed[0]
            # Excel A–M の変数名は従来どおり number,category,level,...
            # INSERT 列は (number, level, category, ...) で、値は従来どおり
            # [number, category, level, ...]（B↔C が入れ替わって DB に入る互換挙動）。
            (
                number,
                category,
                level,
                q,
                a1,
                a2,
                a3,
                a4,
                cid1,
                cid2,
                cid3,
                cid4,
                flag,
            ) = parsed
            c.execute(
                "INSERT INTO knowledge_base "
                "(number, level, category, q, a1, a2, a3, a4, cid1, cid2, cid3, cid4, flag) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [
                    number,
                    category,
                    level,
                    q,
                    a1,
                    a2,
                    a3,
                    a4,
                    cid1,
                    cid2,
                    cid3,
                    cid4,
                    flag,
                ],
            )
    except Exception:
        if number is None:
            number = -1
        conn.rollback()
        conn.close()
        return False, number
    conn.commit()
    conn.close()
    return True, 0

def read_excel2(conn, fname):

    c = conn.cursor()

    wb = load_workbook(fname, data_only=True)
    ws = wb.active  # 1枚目のシート

    try:
        for row in ws.iter_rows(min_row=1, values_only=True):
            comment_id, comment = row
#           print(comment_id, comment)
            c.execute(
                'INSERT INTO comments_table ( comment_id, comment) ' +
                'VALUES (?,?)',
                [comment_id, comment])
    except:
        conn.rollback()
        conn.close()
        return False, -1
    conn.commit()
    conn.close()
    return True, 0

